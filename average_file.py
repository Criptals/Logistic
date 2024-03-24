import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment


def create_average_file(gen_name, avg_name):
    wb = Workbook()
    book = openpyxl.load_workbook(f"{gen_name}.xlsx")
    sheets_num = len(book.sheetnames)
    all_data = []
    form_data = []
    # блок для генерации шапки таблицы средних значений и получения всех необходимых данных из сводной таблицы
    for i in range(sheets_num):
        sheet_data = []
        wa = book.worksheets[i]
        ws = wb.create_sheet(wa.title)  # создание листа
        f_data = []
        for row in wa:
            row_data = []
            for cell in row:
                if int(cell.coordinate[1::]) < 4:
                    ws[cell.coordinate].value = cell.value
                else:
                    if len(f_data) < len(row):
                        f_data.append(cell.number_format)
                    if cell.value is None:
                        row_data.append(0)
                    elif len(str(cell.value)) == str(cell.value).count(" "):
                        row_data.append(0)
                    else:
                        row_data.append(cell.value)
            if len(row_data) > 0:
                sheet_data.append(row_data)
        print(f_data)
        if i < 3:
            ws.delete_cols(4, 1)
            ws.delete_cols(5, 1)
            ws.delete_cols(6, 5)
            f_data.pop(3)
            f_data.pop(4)
            f_data.pop(5)
            f_data.pop(5)
            f_data.pop(5)
            f_data.pop(5)
            f_data.pop(5)

            for k in sheet_data:
                k.pop(3)
                k.pop(4)
                k.pop(5)
                k.pop(5)
                k.pop(5)
                k.pop(5)
                k.pop(5)
        elif i == 3:
            ws.delete_cols(7, 7)
            f_data.pop(6)
            f_data.pop(6)
            f_data.pop(6)
            f_data.pop(6)
            f_data.pop(6)
            f_data.pop(6)
            f_data.pop(6)
            for k in sheet_data:
                k.pop(6)
                k.pop(6)
                k.pop(6)
                k.pop(6)
                k.pop(6)
                k.pop(6)
                k.pop(6)
        elif i == 4:
            ws.delete_cols(10, 2)
            f_data.pop(9)
            f_data.pop(9)
            for k in sheet_data:
                k.pop(9)
                k.pop(9)
        all_data.append(sheet_data)
        if len(form_data) < sheets_num:
            form_data.append(f_data)
    del wb['Sheet']
    wb.save(f"{avg_name}.xlsx")

    # блок вычисления всех средних значений
    avg_data = []
    for k in range(len(all_data)):
        sheet_data = all_data[k]
        avg_sheet_data = []
        cities = []
        values = []
        r = len(sheet_data)
        if k < 3:
            for i in range(r):
                nums = 1
                if not cities:
                    cities.append(sheet_data[i][:3])
                    values.append(sheet_data[i][3::])
                elif sheet_data[i][:3] not in cities:
                    cities.append(sheet_data[i][:3])
                    values.append(sheet_data[i][3::])
                if i != r-1:
                    for j in range(i + 1, r):
                        c = sheet_data[j][:3]
                        v = sheet_data[j][3::]
                        if c == cities[-1]:
                            values[-1][0] += v[0]
                            values[-1][1] += v[1]
                            nums += 1
                    values[-1][0] /= nums
                    values[-1][1] /= nums

            for i in range(len(cities)):
                for f in range(len(cities[i])):
                    if cities[i][f] == 0:
                        cities[i][f] = ""
                avg_sheet_data.append((cities[i] + values[i]))
        elif k == 3:
            for i in range(r):
                nums = 1
                if not cities:
                    cities.append(sheet_data[i][:3])
                    values.append(sheet_data[i][3::])
                elif sheet_data[i][:3] not in cities:
                    cities.append(sheet_data[i][:3])
                    values.append(sheet_data[i][3::])

                if i != r - 1:
                    for j in range(i + 1, r):
                        c = sheet_data[j][:3]
                        v = sheet_data[j][3::]
                        if c == cities[-1]:
                            values[-1][0] += v[0]
                            values[-1][1] += v[1]
                            values[-1][2] += v[2]
                            nums += 1
                    values[-1][0] /= nums
                    values[-1][1] /= nums
                    values[-1][2] /= nums

            for i in range(len(cities)):
                for f in range(len(cities[i])):
                    if cities[i][f] == 0:
                        cities[i][f] = ""
                avg_sheet_data.append((cities[i] + values[i]))
        elif k == 4:
            for i in range(r):
                nums = 1
                if not cities:
                    cities.append(sheet_data[i][:4] + sheet_data[i][6:7] + sheet_data[i][8::])
                    values.append(sheet_data[i][4:6] + sheet_data[i][7:8])
                elif sheet_data[i][:2] not in cities:
                    cities.append(sheet_data[i][:4] + sheet_data[i][6:7] + sheet_data[i][8::])
                    values.append(sheet_data[i][4:6] + sheet_data[i][7:8])

                if i != r - 1:
                    for j in range(i + 1, r):
                        c = sheet_data[j][:4] + sheet_data[j][6:7] + sheet_data[j][8::]
                        v = sheet_data[j][4:6] + sheet_data[j][7:8]
                        if i == 0:
                            print(cities[-1], c)
                            print(values[-1], v)
                            print()
                        if c[:4] == cities[-1][:4]:
                            values[-1][0] += v[0]
                            values[-1][1] += v[1]
                            values[-1][2] += v[2]
                            if i == 0:
                                print(values[-1][0], values[-1][1], values[-1][2])
                            nums += 1
                    values[-1][0] /= nums
                    values[-1][1] /= nums
                    values[-1][2] /= nums
                    if i == 0:
                        print(values[-1][0], values[-1][1], values[-1][2])

            for i in range(len(cities)):
                for f in range(len(cities[i])):
                    if cities[i][f] == 0:
                        cities[i][f] = ""
                avg_sheet_data.append((cities[i][:4] + values[i][:2] + cities[i][4:5] + values[i][2::] + cities[i][5::]))
        avg_data.append(avg_sheet_data)

    letters = ['A', "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    wb = openpyxl.load_workbook(f"{avg_name}.xlsx")
    for i in range(sheets_num):
        ws = wb.worksheets[i]
        sheet_data = avg_data[i]
        row_num = 4
        for rows in sheet_data:
            for j in range(len(rows)):
                ws[letters[j] + str(row_num)].number_format = form_data[i][j]
                ws[letters[j] + str(row_num)].value = rows[j]
            row_num += 1
    wb.save(f"{avg_name}.xlsx")

    # блок форматирования ячеек
    wb = openpyxl.load_workbook(f"{avg_name}.xlsx")
    for i in range(sheets_num):
        ws = wb.worksheets[i]
        columns = ws.max_row
        rows = ws.max_column
        for j in range(1, columns + 1):
            for k in range(1, rows + 1):
                ws.cell(j, k).alignment = Alignment(horizontal="center", vertical="top")
        if i < 3:
            ws.merge_cells("A1:E1")
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
        elif i == 3:
            ws.merge_cells("A1:G1")
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
        elif i == 4:
            ws.merge_cells("A1:J1")
            ws.merge_cells("E2:G2")
            ws.merge_cells("H2:I2")
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 80
    wb.save(f"{avg_name}.xlsx")
