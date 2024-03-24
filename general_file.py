import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from average_file import create_average_file


def create_general_file(a, gen_name, avg_name):
    # блок заполнения сводной таблицы данными из первой таблицы
    wb = Workbook()
    book = openpyxl.load_workbook((a[0]))
    sheets_num = len(book.sheetnames)
    row_num = [0 for i in range(sheets_num)]
    for i in range(sheets_num):
        k = 0
        if i < 3:
            k = 8
        elif i == 4:
            k = 5
        elif i == 5:
            k = 6
        wa = book.worksheets[i]
        ws = wb.create_sheet(wa.title) #создание листа
        for row in wa:
            r = 0
            for cell in row:
                if (int(cell.coordinate[1::]) == 1 or int(cell.coordinate[1::]) > k) and r == 0:
                    row_num[i] += 1
                    r = 1
                    if cell.value is None:
                        ws[cell.coordinate[0] + str(row_num[i])].value = cell.value
                    else:
                        if str(cell.value).isdigit():
                            ws[cell.coordinate[0] + str(row_num[i])].number_format = cell.number_format[:23]
                            ws[cell.coordinate[0] + str(row_num[i])].value = cell.value
                        else:
                            ws[cell.coordinate[0] + str(row_num[i])].value = cell.value
                elif int(cell.coordinate[1::]) == 1 or int(cell.coordinate[1::]) > k and r == 1:
                    if cell.value is None:
                        ws[cell.coordinate[0] + str(row_num[i])].value = cell.value
                    else:
                        if str(cell.value).isdigit():
                            ws[cell.coordinate[0] + str(row_num[i])].number_format = cell.number_format[:23]
                            ws[cell.coordinate[0] + str(row_num[i])].value = cell.value

                        else:
                            ws[cell.coordinate[0] + str(row_num[i])].value = cell.value
    del wb['Sheet']
    wb.save(f"{gen_name}.xlsx")

    # блок заполнения сводной таблицы данными из таблиц, начиная со 2
    files_num = len(a)
    for i in range(1, files_num):
        book = openpyxl.load_workbook((a[i]))
        wb = openpyxl.load_workbook(f"{gen_name}.xlsx")
        a = book.sheetnames.index('авто')
        row_num[a] -= 9
        k = 0
        for j in range(sheets_num):
            wa = book.worksheets[j]
            ws = wb.worksheets[j]
            if j < 3:
                k = 10
            elif j == 4:
                k = 7
            elif j == 5:
                k = 8
            for row in wa:
                r = 0
                for cell in row:
                    if int(cell.coordinate[1::]) > k and r == 0:
                        row_num[j] += 1
                        r = 1
                        if cell.value is None:
                            ws[cell.coordinate[0] + str(row_num[j])].value = cell.value
                        else:
                            if str(cell.value).isdigit():
                                ws[cell.coordinate[0] + str(row_num[j])].number_format = cell.number_format[:23]
                                ws[cell.coordinate[0] + str(row_num[j])].value = cell.value
                            else:
                                ws[cell.coordinate[0] + str(row_num[j])].value = cell.value
                    elif int(cell.coordinate[1::]) > k and r == 1:
                        if cell.value is None:
                            ws[cell.coordinate[0] + str(row_num[j])].value = cell.value
                        else:
                            if str(cell.value).isdigit():
                                ws[cell.coordinate[0] + str(row_num[j])].number_format = cell.number_format[:23]
                                ws[cell.coordinate[0] + str(row_num[j])].value = cell.value
                            else:
                                ws[cell.coordinate[0] + str(row_num[j])].value = cell.value
            wb.save(f"{gen_name}.xlsx")
    wb = openpyxl.load_workbook(f"{gen_name}.xlsx")
    del wb["Лист1"]
    wb.save(f"{gen_name}.xlsx")

    # блок форматирования ячеек
    wb = openpyxl.load_workbook(f"{gen_name}.xlsx")
    for i in range(sheets_num-1):
        ws = wb.worksheets[i]
        columns = ws.max_row
        rows = ws.max_column
        for j in range(1, columns + 1):
            for k in range(1, rows + 1):
                ws.cell(j, k).alignment = Alignment(horizontal="center", vertical="top")
        if i < 3:
            ws.merge_cells("A1:L1")
            ws.merge_cells("E2:F2")
            ws.merge_cells("G2:H2")
            ws.merge_cells("I2:J2")
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 90
            ws.column_dimensions['L'].width = 20
        elif i == 3:
            ws.merge_cells("A1:M1")
            ws.merge_cells("G2:H2")
            ws.merge_cells("J2:K2")
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 60
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 10
        elif i == 4:
            ws.merge_cells("A1:L1")
            ws.merge_cells("E2:G2")
            ws.merge_cells("H2:I2")
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 80
    wb.save(f"{gen_name}.xlsx")
    create_average_file(gen_name=gen_name, avg_name=avg_name)
